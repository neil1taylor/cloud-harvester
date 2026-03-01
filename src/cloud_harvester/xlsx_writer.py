"""Generate XLSX files matching classic_analyser's format."""
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from cloud_harvester.schema import ResourceSchema

HEADER_FILL = PatternFill(start_color="FF0F62FE", end_color="FF0F62FE", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def write_xlsx(
    filepath: str,
    data: dict[str, list[dict]],
    schemas: dict[str, ResourceSchema],
    account_info: dict,
) -> None:
    """Write collected data to XLSX file."""
    wb = Workbook()

    # Summary sheet (first sheet)
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, data, schemas, account_info)

    # Resource worksheets
    for resource_key, rows in data.items():
        schema = schemas.get(resource_key)
        if not schema or not rows:
            continue
        ws = wb.create_sheet(title=schema.worksheet_name)
        _write_resource_sheet(ws, schema, rows)

    wb.save(filepath)


def _write_summary(ws, data, schemas, account_info):
    """Write the Summary sheet with account info and resource counts."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    rows = [
        ("Account Name", account_info.get("name", "")),
        ("Account ID", account_info.get("id", "")),
        ("Account Email", account_info.get("email", "")),
        ("Account Owner", account_info.get("owner", "")),
        ("Collection Timestamp", datetime.now(timezone.utc).isoformat()),
        ("", ""),
    ]

    # Resource counts
    for resource_key, schema in schemas.items():
        count = len(data.get(resource_key, []))
        rows.append((schema.worksheet_name, count))

    for row_idx, (key, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)


def _write_resource_sheet(ws, schema: ResourceSchema, rows: list[dict]):
    """Write a resource worksheet with headers and data."""
    columns = schema.columns

    # Header row
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def.header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_def in enumerate(columns, start=1):
            value = row_data.get(col_def.field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=_format_value(value, col_def.data_type))

    # Auto-width
    for col_idx, col_def in enumerate(columns, start=1):
        max_len = len(col_def.header)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)


def _format_value(value, data_type: str):
    """Format a value for XLSX output."""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    if data_type == "boolean":
        if isinstance(value, bool):
            return "Yes" if value else "No"
        return str(value)
    return value
