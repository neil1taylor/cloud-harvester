# tests/test_integration.py
"""End-to-end integration test: authentication -> collection -> XLSX generation."""
import os
import tempfile
from unittest.mock import patch

from openpyxl import load_workbook

from cloud_harvester.harvester import run_harvest


# ---------------------------------------------------------------------------
# Fake collectors that return known data (same signature as real collectors)
# ---------------------------------------------------------------------------

def _fake_vsi_collector(api_key, token, regions):
    return [
        {
            "id": 1, "hostname": "web01", "domain": "test.com",
            "fqdn": "web01.test.com", "primaryIp": "10.0.0.1",
            "backendIp": "10.0.0.2", "maxCpu": 4, "maxMemory": 8192,
            "status": "ACTIVE", "powerState": "RUNNING", "datacenter": "dal13",
            "os": "Ubuntu 22.04", "hourlyBilling": "No",
            "createDate": "2025-01-01", "recurringFee": "150.00",
            "costBasis": "Monthly", "notes": "", "privateNetworkOnly": "No",
            "localDisk": "No", "startCpus": 4, "modifyDate": "2025-02-01",
            "dedicated": "No", "placementGroupId": 0, "tags": "web",
            "diskGb": 100, "networkVlans": "1234",
        },
        {
            "id": 2, "hostname": "db01", "domain": "test.com",
            "fqdn": "db01.test.com", "primaryIp": "10.0.0.3",
            "backendIp": "10.0.0.4", "maxCpu": 8, "maxMemory": 16384,
            "status": "ACTIVE", "powerState": "RUNNING", "datacenter": "dal13",
            "os": "RHEL 8", "hourlyBilling": "No",
            "createDate": "2025-01-15", "recurringFee": "300.00",
            "costBasis": "Monthly", "notes": "database", "privateNetworkOnly": "Yes",
            "localDisk": "No", "startCpus": 8, "modifyDate": "2025-02-15",
            "dedicated": "No", "placementGroupId": 0, "tags": "db",
            "diskGb": 500, "networkVlans": "1234",
        },
    ]


def _fake_vlan_collector(api_key, token, regions):
    return [
        {
            "id": 100, "vlanNumber": 1234, "name": "private",
            "networkSpace": "PRIVATE", "primaryRouter": "bcr01a.dal13",
            "datacenter": "dal13", "subnetCount": 5,
            "firewallComponents": 0, "gateway": "",
        },
    ]


def _fake_vpc_instance_collector(api_key, token, regions):
    return [
        {
            "id": "vpc-inst-001", "name": "vpc-web01", "status": "running",
            "profile": "bx2-2x8", "vcpu": 2, "memory": 8,
            "zone": "us-south-1", "vpcName": "my-vpc",
            "primaryIp": "10.240.0.5", "region": "us-south",
            "created_at": "2025-03-01", "resourceGroup": "default",
        },
    ]


# ---------------------------------------------------------------------------
# Integration tests
# ---------------------------------------------------------------------------

def test_full_harvest_classic_produces_xlsx():
    """End-to-end: auth -> collect classic domain -> verify XLSX output."""
    fake_classic_collectors = [
        ("virtualServers", _fake_vsi_collector),
        ("vlans", _fake_vlan_collector),
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        with patch("cloud_harvester.harvester.authenticate", return_value="fake-token"), \
             patch("cloud_harvester.harvester.get_account_info", return_value={
                 "name": "Test Account", "account_id": "123456",
                 "owner_email": "admin@test.com", "owner": "Admin User",
             }), \
             patch("cloud_harvester.harvester._get_domain_collectors",
                   return_value=fake_classic_collectors):

            run_harvest(
                api_key="fake-key",
                domains=["classic"],
                skip=[],
                accounts=[],
                regions=[],
                output_dir=tmpdir,
                concurrency=2,
                resume=False,
                no_cache=True,
            )

        # Exactly one XLSX should be produced
        xlsx_files = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
        assert len(xlsx_files) == 1, f"Expected 1 XLSX, found {len(xlsx_files)}: {xlsx_files}"

        filepath = os.path.join(tmpdir, xlsx_files[0])
        wb = load_workbook(filepath)

        # -- Summary sheet --
        assert "Summary" in wb.sheetnames
        ws_summary = wb["Summary"]
        assert ws_summary.cell(row=1, column=1).value == "Account Name"
        assert ws_summary.cell(row=1, column=2).value == "Test Account"
        assert ws_summary.cell(row=2, column=1).value == "Account ID"
        assert ws_summary.cell(row=2, column=2).value == "123456"
        assert ws_summary.cell(row=3, column=1).value == "Account Email"
        assert ws_summary.cell(row=3, column=2).value == "admin@test.com"
        assert ws_summary.cell(row=4, column=1).value == "Account Owner"
        assert ws_summary.cell(row=4, column=2).value == "Admin User"

        # -- vVirtualServers sheet --
        assert "vVirtualServers" in wb.sheetnames
        ws_vsi = wb["vVirtualServers"]
        headers = [cell.value for cell in ws_vsi[1]]
        assert "ID" in headers
        assert "Hostname" in headers
        assert "Datacenter" in headers

        # Row 2 = first server (web01)
        assert ws_vsi.cell(row=2, column=headers.index("ID") + 1).value == 1
        assert ws_vsi.cell(row=2, column=headers.index("Hostname") + 1).value == "web01"

        # Row 3 = second server (db01)
        assert ws_vsi.cell(row=3, column=headers.index("ID") + 1).value == 2
        assert ws_vsi.cell(row=3, column=headers.index("Hostname") + 1).value == "db01"

        # Header styling: bold, white text on blue background
        header_cell = ws_vsi.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "FFFFFFFF"  # white text
        assert header_cell.fill.start_color.rgb == "FF0F62FE"  # IBM blue

        # -- vVLANs sheet --
        assert "vVLANs" in wb.sheetnames
        ws_vlans = wb["vVLANs"]
        vlan_headers = [cell.value for cell in ws_vlans[1]]
        assert "ID" in vlan_headers
        assert "VLAN Number" in vlan_headers
        assert ws_vlans.cell(row=2, column=vlan_headers.index("ID") + 1).value == 100
        assert ws_vlans.cell(row=2, column=vlan_headers.index("VLAN Number") + 1).value == 1234

        wb.close()


def test_full_harvest_multi_domain_produces_xlsx():
    """End-to-end: auth -> collect classic + vpc domains -> verify XLSX output."""
    fake_classic_collectors = [
        ("virtualServers", _fake_vsi_collector),
        ("vlans", _fake_vlan_collector),
    ]
    fake_vpc_collectors = [
        ("vpcInstances", _fake_vpc_instance_collector),
    ]

    def mock_get_domain_collectors(domain):
        if domain == "classic":
            return fake_classic_collectors
        elif domain == "vpc":
            return fake_vpc_collectors
        return []

    with tempfile.TemporaryDirectory() as tmpdir:
        with patch("cloud_harvester.harvester.authenticate", return_value="fake-token"), \
             patch("cloud_harvester.harvester.get_account_info", return_value={
                 "name": "Multi Domain Account", "account_id": "789012",
                 "owner_email": "owner@test.com", "owner": "Test Owner",
             }), \
             patch("cloud_harvester.harvester._get_domain_collectors",
                   side_effect=mock_get_domain_collectors):

            run_harvest(
                api_key="fake-key",
                domains=["classic", "vpc"],
                skip=[],
                accounts=[],
                regions=[],
                output_dir=tmpdir,
                concurrency=2,
                resume=False,
                no_cache=True,
            )

        xlsx_files = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
        assert len(xlsx_files) == 1

        filepath = os.path.join(tmpdir, xlsx_files[0])
        wb = load_workbook(filepath)

        # Summary present
        assert "Summary" in wb.sheetnames
        ws_summary = wb["Summary"]
        assert ws_summary.cell(row=1, column=2).value == "Multi Domain Account"

        # Classic sheets present
        assert "vVirtualServers" in wb.sheetnames
        assert "vVLANs" in wb.sheetnames

        # VPC sheet present
        assert "vVpcInstances" in wb.sheetnames
        ws_vpc = wb["vVpcInstances"]
        vpc_headers = [cell.value for cell in ws_vpc[1]]
        assert "ID" in vpc_headers
        assert "Name" in vpc_headers
        assert "VPC" in vpc_headers
        assert ws_vpc.cell(row=2, column=vpc_headers.index("Name") + 1).value == "vpc-web01"
        assert ws_vpc.cell(row=2, column=vpc_headers.index("VPC") + 1).value == "my-vpc"

        # VPC headers also have correct styling
        vpc_header_cell = ws_vpc.cell(row=1, column=1)
        assert vpc_header_cell.font.bold is True
        assert vpc_header_cell.fill.start_color.rgb == "FF0F62FE"

        wb.close()


def test_full_harvest_skip_resource_type():
    """End-to-end: verify that --skip excludes a resource type from the XLSX."""
    fake_classic_collectors = [
        ("virtualServers", _fake_vsi_collector),
        ("vlans", _fake_vlan_collector),
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        with patch("cloud_harvester.harvester.authenticate", return_value="fake-token"), \
             patch("cloud_harvester.harvester.get_account_info", return_value={
                 "name": "Skip Test", "account_id": "999",
                 "owner_email": "skip@test.com", "owner": "Skip Owner",
             }), \
             patch("cloud_harvester.harvester._get_domain_collectors",
                   return_value=fake_classic_collectors):

            run_harvest(
                api_key="fake-key",
                domains=["classic"],
                skip=["vlans"],
                accounts=[],
                regions=[],
                output_dir=tmpdir,
                concurrency=1,
                resume=False,
                no_cache=True,
            )

        xlsx_files = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
        assert len(xlsx_files) == 1

        filepath = os.path.join(tmpdir, xlsx_files[0])
        wb = load_workbook(filepath)

        # virtualServers should be present
        assert "vVirtualServers" in wb.sheetnames

        # vlans was skipped so sheet should NOT be present (empty data = no sheet)
        assert "vVLANs" not in wb.sheetnames

        wb.close()


def test_full_harvest_xlsx_filename_format():
    """Verify the output filename includes account name, ID, and timestamp."""
    fake_collectors = [("virtualServers", _fake_vsi_collector)]

    with tempfile.TemporaryDirectory() as tmpdir:
        with patch("cloud_harvester.harvester.authenticate", return_value="fake-token"), \
             patch("cloud_harvester.harvester.get_account_info", return_value={
                 "name": "Acme Corp", "account_id": "ACME42",
                 "owner_email": "a@acme.com", "owner": "Acme Admin",
             }), \
             patch("cloud_harvester.harvester._get_domain_collectors",
                   return_value=fake_collectors):

            run_harvest(
                api_key="fake-key",
                domains=["classic"],
                skip=[],
                accounts=[],
                regions=[],
                output_dir=tmpdir,
                concurrency=1,
                resume=False,
                no_cache=True,
            )

        xlsx_files = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
        assert len(xlsx_files) == 1
        filename = xlsx_files[0]

        # Filename should contain sanitised account name and account ID
        assert "Acme_Corp" in filename
        assert "ACME42" in filename
        assert filename.endswith(".xlsx")
