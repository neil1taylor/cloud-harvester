import os
import tempfile
from openpyxl import load_workbook
from cloud_harvester.xlsx_writer import write_xlsx
from cloud_harvester.schema import CLASSIC_SCHEMAS


def test_write_xlsx_creates_file():
    data = {
        "virtualServers": [
            {"id": 123, "hostname": "web01", "domain": "example.com", "fqdn": "web01.example.com",
             "primaryIp": "10.0.0.1", "backendIp": "10.0.0.2", "maxCpu": 4, "maxMemory": 8192,
             "status": "Active", "powerState": "Running", "datacenter": "dal13", "os": "Ubuntu 22.04",
             "hourlyBilling": "No", "createDate": "2025-01-01", "recurringFee": "150.00",
             "costBasis": "Monthly", "notes": "", "privateNetworkOnly": "No", "localDisk": "No",
             "startCpus": 4, "modifyDate": "2025-02-01", "dedicated": "No", "placementGroupId": 0,
             "tags": "web,prod", "diskGb": 100, "networkVlans": "1234,5678"},
        ],
    }
    account_info = {
        "name": "Test Account",
        "id": "12345",
        "email": "test@example.com",
        "owner": "Test Owner",
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, "test.xlsx")
        write_xlsx(filepath, data, CLASSIC_SCHEMAS, account_info)
        assert os.path.exists(filepath)

        wb = load_workbook(filepath)
        assert "Summary" in wb.sheetnames
        assert "vVirtualServers" in wb.sheetnames

        ws = wb["vVirtualServers"]
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Hostname" in headers
        assert ws.cell(row=2, column=1).value == 123
        assert ws.cell(row=2, column=2).value == "web01"


def test_write_xlsx_empty_data_still_creates_summary():
    with tempfile.TemporaryDirectory() as tmpdir:
        filepath = os.path.join(tmpdir, "empty.xlsx")
        write_xlsx(filepath, {}, CLASSIC_SCHEMAS, {"name": "Empty", "id": "0"})
        assert os.path.exists(filepath)
        wb = load_workbook(filepath)
        assert "Summary" in wb.sheetnames
